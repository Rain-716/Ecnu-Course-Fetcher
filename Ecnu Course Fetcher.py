# -*- coding: utf-8 -*-
import os
import re
import json
import time
import base64
import pandas as pd
from io import BytesIO
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.utils import get_column_letter

# 先补回 ANTIALIAS，再导入 ddddocr
from PIL import Image
if not hasattr(Image, 'ANTIALIAS'):
    # Pillow>=10.0.0 中移除了 ANTIALIAS，映射到新的枚举
    Image.ANTIALIAS = Image.Resampling.LANCZOS  # type: ignore

import ddddocr  # 现在内部用到 ANTIALIAS 的地方都能正常工作

# 用户配置
USERNAME = ''
PASSWORD = ''
LOGIN_URL = 'https://applicationnewjw.ecnu.edu.cn/eams/stdElectCourse.action'
TARGET_URL = 'https://applicationnewjw.ecnu.edu.cn/eams/stdElectCourse!queryStdCount.action?profileId=6022'
JS_FILE_COUNTS = 'stdElectCourse!queryStdCount.js'
JS_FILE_LESSONS = 'stdElectCourse!data.js'
OUTPUT_EXCEL = 'lesson_overview_with_counts.xlsx'
INTERVAL_SECONDS = 60  # 循环时间

# 初始化 OCR 引擎
ocr = ddddocr.DdddOcr()

def load_js_json(file_path):
    """从 .js 文件中抽取并标准化 JSON 字符串"""
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    content = re.sub(r'<.*?>', '', content)
    content = re.sub(r'/\*.*?\*/', '', content, flags=re.S)
    start = content.find('=') + 1
    semicolon_pos = content.rfind(';')
    end = semicolon_pos if semicolon_pos > start else len(content)
    data_str = content[start:end]
    data_str = re.sub(r'([{,]\s*)([a-zA-Z_]\w*)(\s*:)', r'\1"\2"\3', data_str)
    data_str = re.sub(r"'([^']*)'", r'"\1"', data_str)
    return data_str

def recognize_captcha(driver, img_element):
    """
    将 Selenium 的 <img> 元素画到 canvas 上并以 Base64 返回，
    然后用 ddddocr 识别并返回验证码文本
    """
    img_base64 = driver.execute_async_script("""
        const img = arguments[0];
        const callback = arguments[1];
        const canvas = document.createElement('canvas');
        canvas.width = img.naturalWidth;
        canvas.height = img.naturalHeight;
        const ctx = canvas.getContext('2d');
        ctx.drawImage(img, 0, 0);
        callback(canvas.toDataURL('image/png').split(',')[1]);
    """, img_element)
    img_data = base64.b64decode(img_base64)
    return ocr.classification(img_data)

def run_task():
    driver = webdriver.Chrome()
    try:
        # 登录流程
        driver.get(LOGIN_URL)
        wait = WebDriverWait(driver, 10)
        wait.until(EC.presence_of_element_located((By.NAME, 'username'))).send_keys(USERNAME)
        driver.find_element(By.CSS_SELECTOR, "#normalLoginForm input[type='password']").send_keys(PASSWORD)

        # 识别并填写验证码
        img_elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#normalLoginForm app-verification img")))
        captcha_text = recognize_captcha(driver, img_elem)
        print(f"识别到的验证码：{captcha_text}")
        driver.find_element(By.CSS_SELECTOR, '#normalLoginForm app-verification input').send_keys(captcha_text)

        # 提交登录
        driver.find_element(By.ID, 'submitBtn').click()

        # 进入选课页面并下载 JS
        enter_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.LINK_TEXT, '进入选课>>>>'))
        )
        enter_button.click()
        driver.get(TARGET_URL)
        time.sleep(1)
        driver.get(TARGET_URL)

        with open(JS_FILE_COUNTS, 'w', encoding='utf-8') as f:
            f.write(driver.page_source)
        print(f"页面内容已保存到 {JS_FILE_COUNTS}")
    finally:
        driver.quit()

    # 解析并合并数据，导出 Excel
    lessons = json.loads(load_js_json(JS_FILE_LESSONS))
    counts = json.loads(load_js_json(JS_FILE_COUNTS))

    df_lessons = pd.DataFrame(lessons)
    drop_cols = [
        'arrangeInfo', 'expLessonGroups',
        'code', 'courseId', 'courseTypeId', 'courseTypeCode',
        'textbooks', 'campusCode', 'remark'
    ]
    df_lessons.drop(columns=[c for c in drop_cols if c in df_lessons.columns], inplace=True)

    df_counts = pd.DataFrame.from_dict(counts, orient='index')
    df_counts.index.name = 'id'
    df_counts.reset_index(inplace=True)
    df_counts.rename(columns={'sc': 'current_count', 'lc': 'limit_count'}, inplace=True)
    df_counts['id'] = df_counts['id'].astype(df_lessons['id'].dtype)

    df = df_lessons.merge(df_counts, on='id', how='left')
    df['remain_count'] = df['limit_count'] - df['current_count']

    with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='课程概览')
        wb = writer.book
        ws = writer.sheets['课程概览']
        for idx, col in enumerate(df.columns, 1):
            max_len = len(col)
            for cell in ws[get_column_letter(idx)]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(idx)].width = max_len + 2

    print(f"已成功将课程信息及人数数据导出到 {OUTPUT_EXCEL}")

if __name__ == '__main__':
    while True:
        run_task()
        time.sleep(INTERVAL_SECONDS)